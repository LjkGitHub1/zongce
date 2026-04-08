from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from django.http import HttpResponse, FileResponse
from django.core.files.storage import default_storage
from django.conf import settings
import csv
import json
import os
import zipfile
import tempfile
from io import StringIO, BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from .models import (
    ResearchProject, Monograph, Paper, OtherAcademic,
    TechnologyCompetition, SocialPractice, SocialService, OtherPractice, File
)
from apps.users.models import User


def export_to_csv(data, filename):
    """导出为CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    
    if not data:
        return response
    
    writer = csv.DictWriter(response, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
    
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_user_materials(request):
    """导出用户材料清单"""
    user = request.user
    
    # 获取所有材料
    all_materials = []
    
    # 课题项目
    for item in ResearchProject.objects.filter(student_id=user.student_id):
        all_materials.append({
            '类型': '课题项目',
            '名称': item.project_name,
            '级别': item.level or '',
            '状态': item.get_status_display(),
            '提交时间': item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        })
    
    # 专著
    for item in Monograph.objects.filter(student_id=user.student_id):
        all_materials.append({
            '类型': '专著',
            '名称': item.title,
            '出版社': item.publish or '',
            '状态': item.get_status_display(),
            '提交时间': item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        })
    
    # 论文
    for item in Paper.objects.filter(student_id=user.student_id):
        all_materials.append({
            '类型': '论文',
            '名称': item.title,
            '刊物': item.publication or '',
            '状态': item.get_status_display(),
            '提交时间': item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        })
    
    # 科技竞赛
    for item in TechnologyCompetition.objects.filter(student_id=user.student_id):
        all_materials.append({
            '类型': '科技竞赛',
            '名称': item.competition,
            '级别': item.get_level_display() if item.level else '',
            '状态': item.get_status_display(),
            '提交时间': item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        })
    
    # 其他材料类型类似处理...
    
    filename = f'材料清单_{user.student_id}_{user.name}'
    return export_to_csv(all_materials, filename)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def export_all_materials(request):
    """导出所有材料（管理员）"""
    identity = request.query_params.get('identity')
    status_filter = request.query_params.get('status')
    
    all_materials = []
    
    # 构建查询条件
    filters = {}
    if identity:
        filters['identity'] = identity
    if status_filter:
        filters['status'] = status_filter
    
    # 课题项目
    for item in ResearchProject.objects.filter(**filters):
        all_materials.append({
            '学号': item.student_id,
            '姓名': item.name,
            '身份': item.get_identity_display(),
            '类型': '课题项目',
            '名称': item.project_name,
            '级别': item.level or '',
            '状态': item.get_status_display(),
            '提交时间': item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        })
    
    # 专著
    for item in Monograph.objects.filter(**filters):
        all_materials.append({
            '学号': item.student_id,
            '姓名': item.name,
            '身份': item.get_identity_display(),
            '类型': '专著',
            '名称': item.title,
            '出版社': item.publish or '',
            '状态': item.get_status_display(),
            '提交时间': item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        })
    
    # 论文
    for item in Paper.objects.filter(**filters):
        all_materials.append({
            '学号': item.student_id,
            '姓名': item.name,
            '身份': item.get_identity_display(),
            '类型': '论文',
            '名称': item.title,
            '刊物': item.publication or '',
            '状态': item.get_status_display(),
            '提交时间': item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        })
    
    # 科技竞赛
    for item in TechnologyCompetition.objects.filter(**filters):
        all_materials.append({
            '学号': item.student_id,
            '姓名': item.name,
            '身份': item.get_identity_display(),
            '类型': '科技竞赛',
            '名称': item.competition,
            '级别': item.get_level_display() if item.level else '',
            '状态': item.get_status_display(),
            '提交时间': item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        })
    
    filename = '所有材料清单'
    if identity:
        filename += f'_{identity}'
    if status_filter:
        filename += f'_{status_filter}'
    
    return export_to_csv(all_materials, filename)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def export_statistics(request):
    """导出统计报表"""
    from .views_stats import admin_statistics
    from apps.system.views import audit_statistics
    
    stats_data = admin_statistics(request).data
    audit_data = audit_statistics(request).data
    
    # 转换为CSV格式
    csv_data = [
        {
            '统计项': '材料总数',
            '数值': stats_data['total']
        },
        {
            '统计项': '待审核',
            '数值': audit_data['pending_count']
        },
        {
            '统计项': '已通过',
            '数值': audit_data['approved_count']
        },
        {
            '统计项': '已驳回',
            '数值': audit_data['rejected_count']
        },
        {
            '统计项': '通过率',
            '数值': f"{audit_data['approval_rate']}%"
        },
    ]
    
    return export_to_csv(csv_data, '统计报表')


def sanitize_filename(filename):
    """清理文件名，移除不允许的字符"""
    import re
    # 移除或替换不允许的字符
    filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
    # 限制文件名长度
    if len(filename) > 200:
        filename = filename[:200]
    return filename


def create_excel_file(data, field_order):
    """创建Excel文件"""
    wb = Workbook()
    ws = wb.active
    
    if not data:
        return wb
    
    # 设置表头
    headers = field_order
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 填充数据
    for row_data in data:
        row = []
        for field in field_order:
            row.append(row_data.get(field, ''))
        ws.append(row)
    
    # 设置列宽自适应
    for col_idx, header in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        # 根据内容调整列宽
        max_length = len(str(header))
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                # 计算单元格内容长度（考虑中文字符，每个中文字符算2个字符宽度）
                cell_str = str(cell_value)
                # 简单估算：中文字符占2个宽度，英文数字占1个宽度
                length = 0
                for char in cell_str:
                    if ord(char) > 127:  # 中文字符
                        length += 2
                    else:
                        length += 1
                max_length = max(max_length, length)
        
        # 设置列宽（至少为表头宽度，最多50个字符）
        # Excel列宽单位：1个单位约等于1个字符宽度
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, len(str(header)) + 2), 50)
    
    return wb


def get_material_title(material, material_type):
    """获取材料的标题/名称"""
    if material_type == 'research_project':
        return material.project_name or '课题项目'
    elif material_type == 'monograph':
        return material.title or '专著'
    elif material_type == 'paper':
        return material.title or '论文'
    elif material_type == 'other_academic':
        return '其他学术成果'
    elif material_type == 'technology_competition':
        return material.title or material.competition or '科技竞赛'
    elif material_type == 'social_practice':
        return material.title or '社会实践调研'
    elif material_type == 'social_service':
        return material.title or '服务社会'
    elif material_type == 'other_practice':
        return material.title or '其他实践活动'
    return '材料'


@api_view(['GET'])
@permission_classes([IsAdminUser])
def download_all_by_class(request):
    """一键下载所有班级的材料（按班级分类）"""
    import logging
    logger = logging.getLogger(__name__)
    
    # 创建临时ZIP文件
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
    temp_file.close()
    
    try:
        with zipfile.ZipFile(temp_file.name, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # 材料类型映射
            academic_types = {
                'research_project': (ResearchProject, '课题项目'),
                'monograph': (Monograph, '专著'),
                'paper': (Paper, '论文'),
                'other_academic': (OtherAcademic, '其他学术成果'),
            }
            
            practice_types = {
                'technology_competition': (TechnologyCompetition, '科技竞赛'),
                'social_practice': (SocialPractice, '社会实践调研'),
                'social_service': (SocialService, '服务社会'),
                'other_practice': (OtherPractice, '其他实践活动'),
            }
            
            # 按身份分组处理：先区分本科生和研究生
            identity_map = {
                'undergraduate': '本科生',
                'graduate': '研究生'
            }
            
            for identity_key, identity_display in identity_map.items():
                # 获取该身份的所有学生
                identity_students = User.objects.filter(identity=identity_key, is_staff=False)
                
                # 获取该身份的所有班级
                classes = identity_students.filter(
                    class_name__isnull=False
                ).exclude(
                    class_name=''
                ).values_list('class_name', flat=True).distinct().order_by('class_name')
                
                for class_name in classes:
                    # 获取该班级该身份的所有学生
                    students = User.objects.filter(class_name=class_name, identity=identity_key, is_staff=False)
                    student_ids = list(students.values_list('student_id', flat=True))
                    
                    if not student_ids:
                        continue
                    
                    # 处理学术成果
                    academic_data = []
                    # 定义字段顺序（按表单顺序，学号和姓名在前）
                    academic_field_orders = {
                        'research_project': ['学号', '姓名', '项目名称', '项目级别', '立项时间', '结题时间', '项目经费（万元）', '项目编号', '主持人（姓名+学号）', '参与人（姓名+学号）', '指导教师', '负责人联系电话', '备注'],
                        'monograph': ['学号', '姓名', '专著名称', '出版社', '出版时间', '编号', '导师姓名', '撰写字数', '作者及顺序', '备注'],
                        'paper': ['学号', '姓名', '论文题目', '刊物名称', '类别', '刊号', '发表时间', '期刊收录情况', '中科院分区', '导师姓名', '是否第一单位', '导师是否通讯', '作者及排名', '备注'],
                        'other_academic': ['学号', '姓名', '成果内容', '备注'],
                    }
                    
                    for type_key, (model, type_name) in academic_types.items():
                        materials = model.objects.filter(student_id__in=student_ids)
                        for material in materials:
                            try:
                                # 获取材料的所有字段（学号和姓名在前）
                                material_dict = {
                                    '学号': material.student_id or '',
                                    '姓名': material.name or '',
                                }
                                
                                # 添加特定字段（按表单顺序）
                                if type_key == 'research_project':
                                    material_dict['项目名称'] = material.project_name or ''
                                    material_dict['项目级别'] = material.get_level_display() if material.level else (material.level or '')
                                    material_dict['立项时间'] = material.start_date.strftime('%Y-%m-%d') if material.start_date else ''
                                    material_dict['结题时间'] = material.end_date.strftime('%Y-%m-%d') if material.end_date else ''
                                    material_dict['项目经费（万元）'] = str(material.funds) if material.funds else ''
                                    material_dict['项目编号'] = material.number or ''
                                    material_dict['主持人（姓名+学号）'] = material.host or ''
                                    material_dict['参与人（姓名+学号）'] = material.participants or ''
                                    material_dict['指导教师'] = material.supervisor or ''
                                    material_dict['负责人联系电话'] = material.contact_phone or ''
                                    material_dict['备注'] = material.remark or ''
                                elif type_key == 'monograph':
                                    material_dict['专著名称'] = material.title or ''
                                    material_dict['出版社'] = material.publish or ''
                                    material_dict['出版时间'] = material.publish_date.strftime('%Y-%m-%d') if material.publish_date else ''
                                    material_dict['编号'] = material.number or ''
                                    material_dict['导师姓名'] = material.supervisor or ''
                                    material_dict['撰写字数'] = str(material.word_count) if material.word_count else ''
                                    material_dict['作者及顺序'] = material.author or ''
                                    material_dict['备注'] = material.remark or ''
                                elif type_key == 'paper':
                                    material_dict['论文题目'] = material.title or ''
                                    material_dict['刊物名称'] = material.publication or ''
                                    material_dict['类别'] = material.category or ''
                                    material_dict['刊号'] = material.publication_number or ''
                                    material_dict['发表时间'] = material.publication_date.strftime('%Y-%m-%d') if material.publication_date else ''
                                    material_dict['期刊收录情况'] = material.journal_status or ''
                                    material_dict['中科院分区'] = material.partition or ''
                                    material_dict['导师姓名'] = material.supervisor or ''
                                    material_dict['是否第一单位'] = '是' if material.first_unit else '否'
                                    material_dict['导师是否通讯'] = '是' if material.communication else '否'
                                    material_dict['作者及排名'] = material.author or ''
                                    material_dict['备注'] = material.remark or ''
                                elif type_key == 'other_academic':
                                    material_dict['成果内容'] = material.content or ''
                                    # OtherAcademic模型没有remark字段
                                    if hasattr(material, 'remark'):
                                        material_dict['备注'] = material.remark or ''
                                    else:
                                        material_dict['备注'] = ''
                                
                                academic_data.append((material_dict, type_key))
                            except Exception as e:
                                import traceback
                                logger.warning(f"处理学术成果材料失败: {type_key}, 材料ID: {material.id if hasattr(material, 'id') else 'unknown'}, 错误: {str(e)}")
                                continue
                            
                            # 收集证明材料
                            files = File.objects.filter(material_type=type_key, material_id=material.id)
                            for file_obj in files:
                                try:
                                    # 检查文件路径
                                    if not file_obj.file_path:
                                        continue
                                    
                                    # 读取文件
                                    if default_storage.exists(file_obj.file_path):
                                        with default_storage.open(file_obj.file_path, 'rb') as f:
                                            file_content = f.read()
                                        
                                        if not file_content:
                                            continue
                                        
                                        # 生成文件名：学号-姓名-作品名称-班级-学术成果-时间戳
                                        material_title = sanitize_filename(get_material_title(material, type_key))
                                        timestamp = int(material.created_at.timestamp())
                                        file_ext = file_obj.file_ext or os.path.splitext(file_obj.original_name)[1] if file_obj.original_name else ''
                                        if file_ext and not file_ext.startswith('.'):
                                            file_ext = '.' + file_ext
                                        if not file_ext:
                                            file_ext = os.path.splitext(file_obj.file_path)[1] or ''
                                        
                                        new_filename = f"{material.student_id}-{material.name}-{material_title}-{class_name}-学术成果-{timestamp}{file_ext}"
                                        new_filename = sanitize_filename(new_filename)
                                        
                                        # 添加到ZIP：身份/班级名/学术成果/证明材料/文件名
                                        zip_path = f"{identity_display}/{class_name}/学术成果/证明材料/{new_filename}"
                                        zip_file.writestr(zip_path, file_content)
                                except Exception as e:
                                    import traceback
                                    logger.warning(f"处理文件失败: {file_obj.file_path if file_obj else 'unknown'}, 错误: {str(e)}")
                                    continue
                    
                    # 处理服务与实践
                    practice_data = []
                    # 定义字段顺序（按表单顺序，学号和姓名在前）
                    practice_field_orders = {
                        'technology_competition': ['学号', '姓名', '赛事名称', '是否为白名单赛事', '作品名称', '级别', '获奖时间', '获奖等级', '组织单位名称', '团体/个人', '负责人（姓名+学号）', '队员（姓名+学号）', '指导老师', '奖状编号', '负责人联系电话', '备注'],
                        'social_practice': ['学号', '姓名', '名称', '级别', '时间', '主持人（姓名+学号）', '参与人（姓名+学号）', '具体内容', '负责人联系电话', '备注'],
                        'social_service': ['学号', '姓名', '名称', '级别', '时间', '作者及顺序', '具体内容', '负责人联系电话', '备注'],
                        'other_practice': ['学号', '姓名', '名称', '级别', '时间', '提供证明单位', '参与人', '具体内容', '负责人联系电话', '备注'],
                    }
                    
                    for type_key, (model, type_name) in practice_types.items():
                        materials = model.objects.filter(student_id__in=student_ids)
                        for material in materials:
                            try:
                                # 获取材料的所有字段（学号和姓名在前）
                                material_dict = {
                                    '学号': material.student_id or '',
                                    '姓名': material.name or '',
                                }
                                
                                # 添加特定字段（按表单顺序）
                                if type_key == 'technology_competition':
                                    material_dict['赛事名称'] = material.competition or ''
                                    material_dict['是否为白名单赛事'] = '是' if material.is_whitelist else '否'
                                    material_dict['作品名称'] = material.title or ''
                                    material_dict['级别'] = material.get_level_display() if material.level else ''
                                    material_dict['获奖时间'] = material.award_date or ''
                                    material_dict['获奖等级'] = material.grade or ''
                                    material_dict['组织单位名称'] = material.organization or ''
                                    material_dict['团体/个人'] = material.group or ''
                                    material_dict['负责人（姓名+学号）'] = material.leader or ''
                                    material_dict['队员（姓名+学号）'] = material.members or ''
                                    material_dict['指导老师'] = material.supervisor or ''
                                    material_dict['奖状编号'] = material.certificate_number or ''
                                    material_dict['负责人联系电话'] = material.contact_phone or ''
                                    material_dict['备注'] = material.remark or ''
                                elif type_key == 'social_practice':
                                    material_dict['名称'] = material.title or ''
                                    material_dict['级别'] = material.get_level_display() if material.level else (material.level or '')
                                    material_dict['时间'] = material.date.strftime('%Y-%m-%d') if material.date else ''
                                    material_dict['主持人（姓名+学号）'] = material.host or ''
                                    material_dict['参与人（姓名+学号）'] = material.participants or ''
                                    material_dict['具体内容'] = material.content or ''
                                    material_dict['负责人联系电话'] = material.contact_phone or ''
                                    material_dict['备注'] = material.remark or ''
                                elif type_key == 'social_service':
                                    material_dict['名称'] = material.title or ''
                                    material_dict['级别'] = material.get_level_display() if material.level else (material.level or '')
                                    material_dict['时间'] = material.date.strftime('%Y-%m-%d') if material.date else ''
                                    material_dict['作者及顺序'] = material.author or ''
                                    material_dict['具体内容'] = material.content or ''
                                    material_dict['负责人联系电话'] = material.contact_phone or ''
                                    material_dict['备注'] = material.remark or ''
                                elif type_key == 'other_practice':
                                    material_dict['名称'] = material.title or ''
                                    material_dict['级别'] = material.get_level_display() if material.level else (material.level or '')
                                    material_dict['时间'] = material.date.strftime('%Y-%m-%d') if material.date else ''
                                    material_dict['提供证明单位'] = material.organization or ''
                                    material_dict['参与人'] = material.author or ''
                                    material_dict['具体内容'] = material.content or ''
                                    material_dict['负责人联系电话'] = material.contact_phone or ''
                                    material_dict['备注'] = material.remark or ''
                                
                                practice_data.append((material_dict, type_key))
                            except Exception as e:
                                import traceback
                                logger.warning(f"处理服务与实践材料失败: {type_key}, 材料ID: {material.id if hasattr(material, 'id') else 'unknown'}, 错误: {str(e)}")
                                continue
                            
                            # 收集证明材料
                            files = File.objects.filter(material_type=type_key, material_id=material.id)
                            for file_obj in files:
                                try:
                                    # 检查文件路径
                                    if not file_obj.file_path:
                                        continue
                                    
                                    if default_storage.exists(file_obj.file_path):
                                        with default_storage.open(file_obj.file_path, 'rb') as f:
                                            file_content = f.read()
                                        
                                        if not file_content:
                                            continue
                                        
                                        material_title = sanitize_filename(get_material_title(material, type_key))
                                        timestamp = int(material.created_at.timestamp())
                                        file_ext = file_obj.file_ext or os.path.splitext(file_obj.original_name)[1] if file_obj.original_name else ''
                                        if file_ext and not file_ext.startswith('.'):
                                            file_ext = '.' + file_ext
                                        if not file_ext:
                                            file_ext = os.path.splitext(file_obj.file_path)[1] or ''
                                        
                                        new_filename = f"{material.student_id}-{material.name}-{material_title}-{class_name}-服务与实践-{timestamp}{file_ext}"
                                        new_filename = sanitize_filename(new_filename)
                                        
                                        zip_path = f"{identity_display}/{class_name}/服务与实践/证明材料/{new_filename}"
                                        zip_file.writestr(zip_path, file_content)
                                except Exception as e:
                                    import traceback
                                    logger.warning(f"处理文件失败: {file_obj.file_path if file_obj else 'unknown'}, 错误: {str(e)}")
                                    continue
                    
                    # 生成Excel文件
                    if academic_data:
                        try:
                            # 按类型分组数据
                            data_by_type = {}
                            for material_dict, type_key in academic_data:
                                if type_key not in data_by_type:
                                    data_by_type[type_key] = []
                                data_by_type[type_key].append(material_dict)
                            
                            # 为每种类型创建Excel文件
                            for type_key, data_list in data_by_type.items():
                                if not data_list:
                                    continue
                                
                                # 获取该类型的字段顺序
                                field_order = academic_field_orders.get(type_key, [])
                                if not field_order:
                                    # 如果没有预定义顺序，使用数据中的字段（学号和姓名在前）
                                    all_fields = set()
                                    for item in data_list:
                                        all_fields.update(item.keys())
                                    field_order = ['学号', '姓名'] + sorted([f for f in all_fields if f not in ['学号', '姓名']])
                                
                                # 创建Excel文件
                                wb = create_excel_file(data_list, field_order)
                                
                                # 保存到内存
                                excel_buffer = BytesIO()
                                wb.save(excel_buffer)
                                excel_buffer.seek(0)
                                
                                # 添加到ZIP
                                type_name_map = {
                                    'research_project': '课题项目',
                                    'monograph': '专著',
                                    'paper': '论文',
                                    'other_academic': '其他学术成果',
                                }
                                excel_filename = f"{identity_display}/{class_name}/学术成果/{type_name_map.get(type_key, '材料清单')}.xlsx"
                                zip_file.writestr(excel_filename, excel_buffer.read())
                        except Exception as e:
                            import traceback
                            logger.error(f"生成学术成果Excel失败: {class_name}/{identity_display}, 错误: {str(e)}")
                            logger.error(traceback.format_exc())
                    
                    # 生成Excel文件
                    if practice_data:
                        try:
                            # 按类型分组数据
                            data_by_type = {}
                            for material_dict, type_key in practice_data:
                                if type_key not in data_by_type:
                                    data_by_type[type_key] = []
                                data_by_type[type_key].append(material_dict)
                            
                            # 为每种类型创建Excel文件
                            for type_key, data_list in data_by_type.items():
                                if not data_list:
                                    continue
                                
                                # 获取该类型的字段顺序
                                field_order = practice_field_orders.get(type_key, [])
                                if not field_order:
                                    # 如果没有预定义顺序，使用数据中的字段（学号和姓名在前）
                                    all_fields = set()
                                    for item in data_list:
                                        all_fields.update(item.keys())
                                    field_order = ['学号', '姓名'] + sorted([f for f in all_fields if f not in ['学号', '姓名']])
                                
                                # 创建Excel文件
                                wb = create_excel_file(data_list, field_order)
                                
                                # 保存到内存
                                excel_buffer = BytesIO()
                                wb.save(excel_buffer)
                                excel_buffer.seek(0)
                                
                                # 添加到ZIP
                                type_name_map = {
                                    'technology_competition': '科技竞赛',
                                    'social_practice': '社会实践调研',
                                    'social_service': '服务社会',
                                    'other_practice': '其他实践活动',
                                }
                                excel_filename = f"{identity_display}/{class_name}/服务与实践/{type_name_map.get(type_key, '材料清单')}.xlsx"
                                zip_file.writestr(excel_filename, excel_buffer.read())
                        except Exception as e:
                            import traceback
                            logger.error(f"生成服务与实践Excel失败: {class_name}/{identity_display}, 错误: {str(e)}")
                            logger.error(traceback.format_exc())
                
                # 为该身份的所有科技竞赛数据生成"竞赛获奖情况"文件夹（不区分班级）
                try:
                    # 获取该身份所有学生的科技竞赛数据
                    identity_student_ids = list(identity_students.values_list('student_id', flat=True))
                    if identity_student_ids:
                        tc_materials = TechnologyCompetition.objects.filter(student_id__in=identity_student_ids)
                        
                        if tc_materials.exists():
                            # 准备Excel数据
                            tc_excel_data = []
                            for material in tc_materials:
                                try:
                                    material_dict = {
                                        '学号': material.student_id or '',
                                        '姓名': material.name or '',
                                        '赛事名称': material.competition or '',
                                        '是否为白名单赛事': '是' if material.is_whitelist else '否',
                                        '作品名称': material.title or '',
                                        '级别': material.get_level_display() if material.level else '',
                                        '获奖时间': material.award_date or '',
                                        '获奖等级': material.grade or '',
                                        '组织单位名称': material.organization or '',
                                        '团体/个人': material.group or '',
                                        '负责人（姓名+学号）': material.leader or '',
                                        '队员（姓名+学号）': material.members or '',
                                        '指导老师': material.supervisor or '',
                                        '奖状编号': material.certificate_number or '',
                                        '负责人联系电话': material.contact_phone or '',
                                        '备注': material.remark or '',
                                    }
                                    tc_excel_data.append(material_dict)
                                except Exception as e:
                                    logger.warning(f"处理科技竞赛材料失败，材料ID: {material.id if hasattr(material, 'id') else 'unknown'}, 错误: {str(e)}")
                                    continue
                            
                            # 生成Excel文件
                            if tc_excel_data:
                                field_order = ['学号', '姓名', '赛事名称', '是否为白名单赛事', '作品名称', '级别', '获奖时间', '获奖等级', '组织单位名称', '团体/个人', '负责人（姓名+学号）', '队员（姓名+学号）', '指导老师', '奖状编号', '负责人联系电话', '备注']
                                wb = create_excel_file(tc_excel_data, field_order)
                                excel_buffer = BytesIO()
                                wb.save(excel_buffer)
                                excel_buffer.seek(0)
                                excel_filename = f"{identity_display}/竞赛获奖情况/竞赛获奖情况.xlsx"
                                zip_file.writestr(excel_filename, excel_buffer.read())
                            
                            # 收集证明材料，命名格式：级别-获奖等级-学号-姓名-作品名称-时间戳
                            for material in tc_materials:
                                files = File.objects.filter(material_type='technology_competition', material_id=material.id)
                                for file_obj in files:
                                    try:
                                        if not file_obj.file_path or not default_storage.exists(file_obj.file_path):
                                            continue
                                        
                                        with default_storage.open(file_obj.file_path, 'rb') as f:
                                            file_content = f.read()
                                        
                                        if not file_content:
                                            continue
                                        
                                        # 获取级别和获奖等级
                                        level_display = material.get_level_display() if material.level else '未知级别'
                                        grade = material.grade or '未知等级'
                                        material_title = sanitize_filename(get_material_title(material, 'technology_competition'))
                                        timestamp = int(material.created_at.timestamp())
                                        
                                        # 文件扩展名
                                        file_ext = file_obj.file_ext or os.path.splitext(file_obj.original_name)[1] if file_obj.original_name else ''
                                        if file_ext and not file_ext.startswith('.'):
                                            file_ext = '.' + file_ext
                                        if not file_ext:
                                            file_ext = os.path.splitext(file_obj.file_path)[1] or ''
                                        
                                        # 命名格式：级别-获奖等级-学号-姓名-作品名称-时间戳
                                        new_filename = f"{level_display}-{grade}-{material.student_id}-{material.name}-{material_title}-{timestamp}{file_ext}"
                                        new_filename = sanitize_filename(new_filename)
                                        
                                        # 添加到ZIP：身份/竞赛获奖情况/证明材料/文件名
                                        zip_path = f"{identity_display}/竞赛获奖情况/证明材料/{new_filename}"
                                        zip_file.writestr(zip_path, file_content)
                                    except Exception as e:
                                        import traceback
                                        logger.warning(f"处理科技竞赛证明材料失败: {file_obj.file_path if file_obj else 'unknown'}, 错误: {str(e)}")
                                        continue
                except Exception as e:
                    import traceback
                    logger.error(f"生成竞赛获奖情况失败: {identity_display}, 错误: {str(e)}")
                    logger.error(traceback.format_exc())
        
        # 返回ZIP文件
        try:
            file_handle = open(temp_file.name, 'rb')
            response = FileResponse(
                file_handle,
                content_type='application/zip'
            )
            # 文件名格式：年月日_材料汇总.zip，例如：2024年01月15日_材料汇总.zip
            filename = f'{datetime.now().strftime("%Y年%m月%d日")}_材料汇总.zip'
            # 确保文件名编码正确
            import urllib.parse
            encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
            response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
            return response
        except Exception as e:
            if 'file_handle' in locals():
                file_handle.close()
            raise
        
    except Exception as e:
        import traceback
        import logging
        logger = logging.getLogger(__name__)
        error_msg = f'生成下载文件失败: {str(e)}'
        error_trace = traceback.format_exc()
        logger.error(f"{error_msg}\n{error_trace}")
        print(error_msg)
        print(error_trace)
        # 返回详细的错误信息用于调试
        return Response({
            'error': error_msg,
            'detail': str(e),
            'traceback': error_trace if settings.DEBUG else None
        }, status=500)
    finally:
        # 清理临时文件
        if os.path.exists(temp_file.name):
            try:
                os.unlink(temp_file.name)
            except:
                pass

