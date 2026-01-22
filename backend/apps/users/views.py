from rest_framework import status, generics, permissions, viewsets
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.response import Response
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from django.contrib.auth import get_user_model
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import load_workbook
import io
from .serializers import UserSerializer, UserRegisterSerializer, UserProfileSerializer, UserAdminSerializer

User = get_user_model()


class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):
    """自定义Token序列化器"""
    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)
        token['student_id'] = user.student_id
        token['name'] = user.name
        token['identity'] = user.identity
        return token


class CustomTokenObtainPairView(TokenObtainPairView):
    """自定义Token获取视图"""
    serializer_class = CustomTokenObtainPairSerializer
    
    def post(self, request, *args, **kwargs):
        """登录并记录日志"""
        response = super().post(request, *args, **kwargs)
        
        # 登录成功后记录日志
        if response.status_code == 200:
            try:
                from apps.system.models import OperationLog
                # 从请求数据中获取学号
                student_id = request.data.get('student_id')
                if student_id:
                    user = User.objects.filter(student_id=student_id).first()
                    if user:
                        ip_address = self._get_client_ip(request)
                        user_agent = request.META.get('HTTP_USER_AGENT', '')[:500]
                        
                        OperationLog.objects.create(
                            user=user,
                            user_name=user.name if hasattr(user, 'name') else user.student_id,
                            operation='登录',
                            module='用户管理',
                            content='用户登录系统',
                            ip_address=ip_address,
                            user_agent=user_agent
                        )
            except Exception as e:
                # 记录日志失败不影响登录
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'记录登录日志失败: {str(e)}')
        
        return response
    
    def _get_client_ip(self, request):
        """获取客户端IP地址"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip or ''


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def register(request):
    """用户注册"""
    serializer = UserRegisterSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.save()
        return Response({
            'message': '注册成功',
            'user': UserSerializer(user).data
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UserProfileView(generics.RetrieveUpdateAPIView):
    """用户个人信息视图"""
    serializer_class = UserProfileSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_object(self):
        return self.request.user


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def change_password(request):
    """修改密码"""
    old_password = request.data.get('old_password')
    new_password = request.data.get('new_password')
    new_password_confirm = request.data.get('new_password_confirm')
    
    if not all([old_password, new_password, new_password_confirm]):
        return Response({'error': '缺少必要参数'}, status=status.HTTP_400_BAD_REQUEST)
    
    if new_password != new_password_confirm:
        return Response({'error': '两次新密码输入不一致'}, status=status.HTTP_400_BAD_REQUEST)
    
    user = request.user
    if not user.check_password(old_password):
        return Response({'error': '原密码错误'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        from django.contrib.auth.password_validation import validate_password
        validate_password(new_password, user)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    user.set_password(new_password)
    user.save()
    
    return Response({'message': '密码修改成功'}, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def user_info(request):
    """获取当前用户信息"""
    serializer = UserSerializer(request.user)
    return Response(serializer.data)


class UserViewSet(viewsets.ModelViewSet):
    """用户管理视图集"""
    queryset = User.objects.all()
    serializer_class = UserAdminSerializer
    permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['identity', 'is_staff', 'is_active', 'major', 'class_name']
    search_fields = ['student_id', 'name', 'major', 'class_name']
    ordering_fields = ['created_at', 'student_id']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'list' or self.action == 'retrieve':
            return UserSerializer
        return UserAdminSerializer
    
    @action(detail=False, methods=['post'])
    def bulk_import(self, request):
        """批量导入用户"""
        if 'file' not in request.FILES:
            return Response({'error': '请上传Excel文件'}, status=status.HTTP_400_BAD_REQUEST)
        
        file = request.FILES['file']
        default_password = 'Aa@2026'
        
        try:
            # 读取Excel文件
            wb = load_workbook(io.BytesIO(file.read()))
            ws = wb.active
            
            # 获取表头
            headers = [cell.value for cell in ws[1]]
            
            # 定义字段映射
            field_map = {
                '学号': 'student_id',
                '姓名': 'name',
                '身份': 'identity',
                '专业': 'major',
                '班级': 'class_name',
                '邮箱': 'email',
                '手机号': 'phone'
            }
            
            # 创建列索引映射
            col_map = {}
            for idx, header in enumerate(headers, start=1):
                if header in field_map:
                    col_map[field_map[header]] = idx
            
            if 'student_id' not in col_map or 'name' not in col_map or 'identity' not in col_map:
                return Response({
                    'error': 'Excel文件必须包含：学号、姓名、身份列'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # 处理数据
            success_count = 0
            error_count = 0
            errors = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                # 跳过空行
                if not row[col_map['student_id'] - 1].value:
                    continue
                
                try:
                    user_data = {}
                    for field, col_idx in col_map.items():
                        cell_value = row[col_idx - 1].value
                        if cell_value:
                            if field == 'identity':
                                # 转换身份
                                identity_map = {
                                    '老师': 'teacher',
                                    '研究生': 'graduate',
                                    '本科生': 'undergraduate',
                                    'teacher': 'teacher',
                                    'graduate': 'graduate',
                                    'undergraduate': 'undergraduate'
                                }
                                cell_str = str(cell_value).strip()
                                user_data[field] = identity_map.get(cell_str, cell_str)
                            else:
                                user_data[field] = str(cell_value) if cell_value else None
                    
                    # 检查必填字段
                    if not user_data.get('student_id') or not user_data.get('name') or not user_data.get('identity'):
                        errors.append(f'第{row_idx}行：缺少必填字段')
                        error_count += 1
                        continue
                    
                    # 检查用户是否已存在
                    if User.objects.filter(student_id=user_data['student_id']).exists():
                        errors.append(f'第{row_idx}行：学号 {user_data["student_id"]} 已存在')
                        error_count += 1
                        continue
                    
                    # 创建用户
                    User.objects.create_user(
                        password=default_password,
                        **user_data
                    )
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f'第{row_idx}行：{str(e)}')
                    error_count += 1
            
            return Response({
                'message': f'导入完成：成功 {success_count} 条，失败 {error_count} 条',
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors[:10]  # 只返回前10个错误
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'error': f'文件处理失败：{str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

